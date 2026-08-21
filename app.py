import streamlit as st
import pandas as pd
import re

from datetime import datetime, timedelta, time
from zoneinfo import ZoneInfo
from io import BytesIO

from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter


# ============================================================
# CONFIGURACIÓN DE LA PÁGINA
# ============================================================

st.set_page_config(
    page_title="Analizador de WhatsApp",
    page_icon="📊",
    layout="wide"
)


# ============================================================
# AGENTES / CHATS
# ============================================================

AGENTES = [
    "Centro",
    "Occidente",
    "Venpro",
    "GranPro",
    "Posmgt25",
    "Oriente",
    "TPOS",
    "POSMaracay",
    "Virtualnet",
    "Tipo II",
    "Multitienda"
]


# ============================================================
# REMITENTES DEL EQUIPO
# ============================================================

NUMEROS_EQUIPO = {
    "584242953004",
    "584166109623",
    "584223304698",
    "584241917586",
    "584125840961",
    "584129920211",
    "584123965504"
}

NOMBRES_EQUIPO = {
    "agentesautorizados",
    "rubenccr"
}


# ============================================================
# FUNCIONES GENERALES
# ============================================================

def normalizar_texto(texto):

    texto = texto.replace("\u202f", " ")
    texto = texto.replace("\xa0", " ")

    return texto


def normalizar_remitente(remitente):

    remitente = remitente.strip()

    numero = re.sub(
        r"\D",
        "",
        remitente
    )

    nombre = re.sub(
        r"\s+",
        "",
        remitente.lower()
    )

    return numero, nombre


def es_remitente_equipo(remitente):

    numero, nombre = normalizar_remitente(
        remitente
    )

    if numero in NUMEROS_EQUIPO:
        return True

    if nombre in NOMBRES_EQUIPO:
        return True

    return False


# ============================================================
# IDENTIFICAR AGENTE SEGÚN NOMBRE DEL ARCHIVO
# ============================================================

def identificar_agente(nombre_archivo):
    """
    Intenta detectar automáticamente a qué agente
    pertenece cada TXT utilizando el nombre del archivo.

    Ejemplos:
    Chat de WhatsApp con Centro.txt
    Centro.txt
    chat_centro_19-08.txt

    -> Centro
    """

    nombre = nombre_archivo.lower()

    nombre_limpio = re.sub(
        r"[^a-z0-9]",
        "",
        nombre
    )

    equivalencias = {

        "Centro": [
            "centro"
        ],

        "Occidente": [
            "occidente"
        ],

        "Venpro": [
            "venpro"
        ],

        "GranPro": [
            "granpro"
        ],

        "Posmgt25": [
            "posmgt25",
            "posmg25"
        ],

        "Oriente": [
            "oriente"
        ],

        "TPOS": [
            "tpos"
        ],

        "POSMaracay": [
            "posmaracay"
        ],

        "Virtualnet": [
            "virtualnet"
        ],

        "Tipo II": [
            "tipoii",
            "tipo2"
        ],

        "Multitienda": [
            "multitienda"
        ]
    }

    for agente, palabras in equivalencias.items():

        for palabra in palabras:

            palabra_limpia = re.sub(
                r"[^a-z0-9]",
                "",
                palabra.lower()
            )

            if palabra_limpia in nombre_limpio:
                return agente

    # Si no logra reconocerlo,
    # utiliza el nombre del archivo sin .txt.

    nombre_base = re.sub(
        r"\.txt$",
        "",
        nombre_archivo,
        flags=re.IGNORECASE
    )

    nombre_base = re.sub(
        r"^chat de whatsapp con\s*",
        "",
        nombre_base,
        flags=re.IGNORECASE
    )

    return nombre_base.strip()


# ============================================================
# FECHA Y HORA
# ============================================================

def convertir_fecha_hora(
    fecha_texto,
    hora_texto,
    periodo
):

    periodo = periodo.lower().strip()

    hora = datetime.strptime(
        hora_texto,
        "%H:%M"
    )

    horas = hora.hour
    minutos = hora.minute

    if "p" in periodo and horas != 12:
        horas += 12

    if "a" in periodo and horas == 12:
        horas = 0

    fecha = datetime.strptime(
        fecha_texto,
        "%d/%m/%Y"
    )

    return datetime(
        fecha.year,
        fecha.month,
        fecha.day,
        horas,
        minutos
    )


# ============================================================
# LEER CHAT WHATSAPP
# ============================================================

def leer_chat_whatsapp(contenido):

    contenido = normalizar_texto(
        contenido
    )

    lineas = contenido.splitlines()

    mensajes = []

    mensaje_actual = None

    patron = re.compile(
        r"^(\d{1,2}/\d{1,2}/\d{4}),\s*"
        r"(\d{1,2}:\d{2})\s*"
        r"([ap]\.\s*m\.)\s*-\s*"
        r"([^:]+):\s*(.*)$",
        re.IGNORECASE
    )

    for linea in lineas:

        coincidencia = patron.match(
            linea
        )

        if coincidencia:

            if mensaje_actual is not None:

                mensajes.append(
                    mensaje_actual
                )

            fecha_texto = coincidencia.group(1)
            hora_texto = coincidencia.group(2)
            periodo = coincidencia.group(3)
            remitente = coincidencia.group(4).strip()
            texto = coincidencia.group(5).strip()

            try:

                fecha_hora = convertir_fecha_hora(
                    fecha_texto,
                    hora_texto,
                    periodo
                )

            except Exception:

                fecha_hora = None

            mensaje_actual = {

                "fecha_hora":
                    fecha_hora,

                "remitente":
                    remitente,

                "mensaje":
                    texto
            }

        else:

            if mensaje_actual is not None:

                mensaje_actual["mensaje"] += (
                    "\n" + linea
                )

    if mensaje_actual is not None:

        mensajes.append(
            mensaje_actual
        )

    return mensajes


# ============================================================
# DETECTAR RIF O SERIAL EN EL MENSAJE
# ============================================================

def mensaje_tiene_rif_o_serial(mensaje):

    tiene_rif = re.search(
        r"\bRIF\s*:",
        mensaje,
        re.IGNORECASE
    )

    tiene_serial = re.search(
        r"\bSERIAL\s*:",
        mensaje,
        re.IGNORECASE
    )

    return bool(
        tiene_rif or tiene_serial
    )


# ============================================================
# DETECTAR SOLICITUD-R AL INICIO
# ============================================================

def empieza_por_solicitud_r(mensaje):

    texto = mensaje.strip()

    return bool(
        re.match(
            r"^[\s*_~]*"
            r"SOLICITUD\s*-\s*R"
            r"\s*:?"
            r"[\s*_~]*",
            texto,
            re.IGNORECASE
        )
    )


# ============================================================
# CLASIFICAR MENSAJE
# ============================================================

def detectar_tipo_solicitud(
    mensaje,
    remitente
):

    es_equipo = es_remitente_equipo(
        remitente
    )

    tiene_identificador = (
        mensaje_tiene_rif_o_serial(
            mensaje
        )
    )

    inicia_solicitud_r = (
        empieza_por_solicitud_r(
            mensaje
        )
    )


    # --------------------------------------------------------
    # EQUIPO
    # --------------------------------------------------------

    if es_equipo:

        if tiene_identificador:
            return "SOLICITUD-R"

        if inicia_solicitud_r:
            return "SOLICITUD-R"

        return None


    # --------------------------------------------------------
    # CLIENTE
    # --------------------------------------------------------

    else:

        if tiene_identificador:
            return "SOLICITUD"

        return None


# ============================================================
# EXTRAER RIF
# ============================================================

def extraer_rif(mensaje):

    patron = re.search(
        r"RIF\s*:\s*([A-Z]?\s*-?\s*\d+)",
        mensaje,
        re.IGNORECASE
    )

    if not patron:
        return ""

    rif = patron.group(1)

    rif = rif.upper()
    rif = rif.replace(" ", "")
    rif = rif.replace("-", "")

    return rif


# ============================================================
# ANALIZAR MENSAJES
# ============================================================

def analizar_solicitudes(
    mensajes,
    desde,
    hasta
):

    registros = []

    for mensaje in mensajes:

        fecha_hora = mensaje[
            "fecha_hora"
        ]

        if fecha_hora is None:
            continue

        # Aplicar MISMO rango a todos los chats
        if not (
            desde <= fecha_hora <= hasta
        ):
            continue

        tipo = detectar_tipo_solicitud(
            mensaje["mensaje"],
            mensaje["remitente"]
        )

        if tipo is None:
            continue

        rif = extraer_rif(
            mensaje["mensaje"]
        )

        registros.append({

            "Fecha y hora":
                fecha_hora,

            "Remitente":
                mensaje["remitente"],

            "Tipo":
                tipo,

            "RIF":
                rif,

            "Mensaje":
                mensaje["mensaje"]
        })

    return pd.DataFrame(
        registros
    )


# ============================================================
# RELACIONAR SOLICITUD CON RESPUESTA
# ============================================================

def relacionar_solicitudes(df):

    solicitudes = df[
        df["Tipo"] == "SOLICITUD"
    ].copy()

    respuestas = df[
        df["Tipo"] == "SOLICITUD-R"
    ].copy()

    solicitudes = solicitudes.sort_values(
        "Fecha y hora"
    )

    respuestas = respuestas.sort_values(
        "Fecha y hora"
    )

    respuestas_usadas = set()

    resultado = []


    for indice_solicitud, solicitud in solicitudes.iterrows():

        rif = solicitud[
            "RIF"
        ]

        respuesta_encontrada = None

        indice_respuesta_encontrada = None


        # ----------------------------------------------------
        # MATCH POR RIF
        # ----------------------------------------------------

        if rif:

            for indice_respuesta, respuesta in respuestas.iterrows():

                if indice_respuesta in respuestas_usadas:
                    continue

                if respuesta["RIF"] != rif:
                    continue

                if (
                    respuesta["Fecha y hora"]
                    < solicitud["Fecha y hora"]
                ):
                    continue

                respuesta_encontrada = (
                    respuesta
                )

                indice_respuesta_encontrada = (
                    indice_respuesta
                )

                break


        # ----------------------------------------------------
        # CONTESTADA
        # ----------------------------------------------------

        if respuesta_encontrada is not None:

            respuestas_usadas.add(
                indice_respuesta_encontrada
            )

            diferencia = (
                respuesta_encontrada[
                    "Fecha y hora"
                ]
                - solicitud[
                    "Fecha y hora"
                ]
            )

            minutos_respuesta = (
                diferencia.total_seconds()
                / 60
            )

            estado = "Contestada"

            fecha_respuesta = (
                respuesta_encontrada[
                    "Fecha y hora"
                ]
            )


        # ----------------------------------------------------
        # SIN CONTESTAR
        # ----------------------------------------------------

        else:

            estado = "Sin contestar"

            fecha_respuesta = None

            minutos_respuesta = None


        resultado.append({

            "Fecha solicitud":
                solicitud[
                    "Fecha y hora"
                ],

            "Solicitante":
                solicitud[
                    "Remitente"
                ],

            "RIF":
                rif,

            "Estado":
                estado,

            "Fecha respuesta":
                fecha_respuesta,

            "Tiempo respuesta (min)":
                round(
                    minutos_respuesta,
                    1
                )
                if minutos_respuesta is not None
                else None
        })


    return pd.DataFrame(
        resultado
    )


# ============================================================
# FORMATEAR TIEMPO
# ============================================================

def formatear_tiempo(minutos):

    if minutos is None:
        return "—"

    if pd.isna(minutos):
        return "—"

    minutos = round(
        float(minutos)
    )

    if minutos < 60:

        return (
            f"{minutos} min"
        )

    horas = (
        minutos // 60
    )

    minutos_restantes = (
        minutos % 60
    )

    if minutos_restantes == 0:

        return (
            f"{horas} h"
        )

    return (
        f"{horas} h "
        f"{minutos_restantes} min"
    )


# ============================================================
# ESTILIZAR EXCEL
# ============================================================

def estilizar_excel(writer):

    color_encabezado = "D9E5F2"

    borde_fino = Side(
        style="thin",
        color="808080"
    )

    for nombre_hoja in writer.book.sheetnames:

        hoja = writer.book[
            nombre_hoja
        ]

        hoja.freeze_panes = "A2"

        # ----------------------------------------------------
        # ENCABEZADOS
        # ----------------------------------------------------

        for celda in hoja[1]:

            celda.font = Font(
                bold=True
            )

            celda.fill = PatternFill(
                fill_type="solid",
                fgColor=color_encabezado
            )

            celda.alignment = Alignment(
                vertical="center"
            )

            celda.border = Border(
                left=borde_fino,
                right=borde_fino,
                top=borde_fino,
                bottom=borde_fino
            )


        # ----------------------------------------------------
        # TODAS LAS CELDAS
        # ----------------------------------------------------

        for fila in hoja.iter_rows():

            for celda in fila:

                celda.border = Border(
                    left=borde_fino,
                    right=borde_fino,
                    top=borde_fino,
                    bottom=borde_fino
                )

                celda.alignment = Alignment(
                    vertical="center"
                )


        # ----------------------------------------------------
        # ALTO DE FILAS
        # ----------------------------------------------------

        hoja.row_dimensions[1].height = 25

        for numero_fila in range(
            2,
            hoja.max_row + 1
        ):

            hoja.row_dimensions[
                numero_fila
            ].height = 22


        # ----------------------------------------------------
        # ANCHO DE COLUMNAS
        # ----------------------------------------------------

        if nombre_hoja == "Resumen":

            anchos = {

                "A": 20,   # Agente

                "B": 15,   # Solicitudes

                "C": 15,   # Contestados

                "D": 17,   # Sin contestar

                "E": 20,   # Tasa

                "F": 32    # Tiempo promedio
            }


        elif nombre_hoja == "Solicitudes":

            anchos = {

                "A": 20,   # Agente

                "B": 23,   # Fecha solicitud

                "C": 24,   # Solicitante

                "D": 18,   # RIF

                "E": 17,   # Estado

                "F": 23,   # Fecha respuesta

                "G": 24,   # Tiempo

                "H": 40    # Observaciones
            }


        else:

            anchos = {}


        for columna, ancho in anchos.items():

            hoja.column_dimensions[
                columna
            ].width = ancho


        # ----------------------------------------------------
        # FORMATO DE FECHAS
        # ----------------------------------------------------

        if nombre_hoja == "Solicitudes":

            for fila in range(
                2,
                hoja.max_row + 1
            ):

                hoja[
                    f"B{fila}"
                ].number_format = (
                    "dd/mm/yyyy hh:mm AM/PM"
                )

                hoja[
                    f"F{fila}"
                ].number_format = (
                    "dd/mm/yyyy hh:mm AM/PM"
                )


# ============================================================
# CREAR EXCEL
# ============================================================

def crear_excel(
    resumen,
    detalle
):

    salida = BytesIO()

    with pd.ExcelWriter(
        salida,
        engine="openpyxl"
    ) as writer:

        # ----------------------------------------------------
        # HOJA 1 - RESUMEN
        # ----------------------------------------------------

        resumen.to_excel(
            writer,
            index=False,
            sheet_name="Resumen"
        )


        # ----------------------------------------------------
        # HOJA 2 - SOLICITUDES
        # ----------------------------------------------------

        detalle.to_excel(
            writer,
            index=False,
            sheet_name="Solicitudes"
        )


        # ----------------------------------------------------
        # FORMATO
        # ----------------------------------------------------

        estilizar_excel(
            writer
        )


    salida.seek(0)

    return salida


# ============================================================
# INTERFAZ
# ============================================================

st.title(
    "Analizador de Atención al Cliente"
)

st.write(
    "Carga los archivos TXT exportados de WhatsApp "
    "para generar las estadísticas."
)


# ============================================================
# CARGAR VARIOS ARCHIVOS
# ============================================================

archivos = st.file_uploader(
    "Sube los archivos TXT exportados de WhatsApp",
    type=["txt"],
    accept_multiple_files=True
)


# ============================================================
# SI SE CARGARON ARCHIVOS
# ============================================================

if archivos:

    cantidad_archivos = len(
        archivos
    )

    st.success(
        f"{cantidad_archivos} archivo(s) cargado(s) correctamente"
    )

    st.divider()


    # ========================================================
    # PERÍODO
    # ========================================================

    st.subheader(
        "Período a analizar"
    )


    zona_venezuela = ZoneInfo(
        "America/Caracas"
    )

    ahora = datetime.now(
        zona_venezuela
    )

    hoy = ahora.date()


    # --------------------------------------------------------
    # LUNES -> VIERNES
    # --------------------------------------------------------

    if hoy.weekday() == 0:

        fecha_desde_default = (
            hoy
            - timedelta(days=3)
        )

    else:

        fecha_desde_default = (
            hoy
            - timedelta(days=1)
        )


    # --------------------------------------------------------
    # 4:30 PM
    # --------------------------------------------------------

    hora_default = time(
        16,
        30
    )


    # ========================================================
    # INTERVALO
    # ========================================================

    intervalo_minutos = st.selectbox(

        "Intervalo para seleccionar la hora",

        options=[
            5,
            10,
            15,
            20,
            30
        ],

        index=1,

        format_func=lambda x:
            f"Cada {x} minutos"
    )


    # ========================================================
    # LISTA DE HORAS
    # ========================================================

    horas = []

    for hora in range(24):

        for minuto in range(
            0,
            60,
            intervalo_minutos
        ):

            hora_obj = time(
                hora,
                minuto
            )

            texto_hora = datetime.strptime(
                f"{hora:02d}:{minuto:02d}",
                "%H:%M"
            ).strftime(
                "%I:%M %p"
            )

            texto_hora = (
                texto_hora.lstrip("0")
            )

            horas.append(
                (
                    texto_hora,
                    hora_obj
                )
            )


    def minutos_del_dia(
        hora_obj
    ):

        return (
            hora_obj.hour * 60
            + hora_obj.minute
        )


    minutos_default = minutos_del_dia(
        hora_default
    )


    indice_default = min(

        range(
            len(horas)
        ),

        key=lambda i: abs(

            minutos_del_dia(
                horas[i][1]
            )

            - minutos_default
        )
    )


    # ========================================================
    # DESDE / HASTA
    # ========================================================

    col1, col2 = st.columns(
        2
    )


    with col1:

        st.markdown(
            "### Desde"
        )

        fecha_desde = st.date_input(

            "Fecha desde",

            value=
                fecha_desde_default,

            format=
                "DD/MM/YYYY",

            key=
                "fecha_desde"
        )

        hora_desde_texto = st.selectbox(

            "Hora desde",

            options=[
                texto
                for texto, _ in horas
            ],

            index=
                indice_default,

            key=
                "hora_desde"
        )

        hora_desde = dict(
            horas
        )[
            hora_desde_texto
        ]


    with col2:

        st.markdown(
            "### Hasta"
        )

        fecha_hasta = st.date_input(

            "Fecha hasta",

            value=
                hoy,

            format=
                "DD/MM/YYYY",

            key=
                "fecha_hasta"
        )

        hora_hasta_texto = st.selectbox(

            "Hora hasta",

            options=[
                texto
                for texto, _ in horas
            ],

            index=
                indice_default,

            key=
                "hora_hasta"
        )

        hora_hasta = dict(
            horas
        )[
            hora_hasta_texto
        ]


    # ========================================================
    # COMBINAR
    # ========================================================

    desde = datetime.combine(
        fecha_desde,
        hora_desde
    )

    hasta = datetime.combine(
        fecha_hasta,
        hora_hasta
    )

    st.divider()


    # ========================================================
    # VALIDACIÓN
    # ========================================================

    if desde >= hasta:

        st.error(
            "La fecha y hora 'Desde' debe ser anterior "
            "a la fecha y hora 'Hasta'."
        )


    else:

        desde_texto = desde.strftime(
            "%d/%m/%Y %I:%M %p"
        )

        hasta_texto = hasta.strftime(
            "%d/%m/%Y %I:%M %p"
        )

        desde_texto = desde_texto.replace(
            " 0",
            " "
        )

        hasta_texto = hasta_texto.replace(
            " 0",
            " "
        )


        st.info(
            f"Período seleccionado para los "
            f"{cantidad_archivos} chats: "
            f"desde {desde_texto} "
            f"hasta {hasta_texto}"
        )


        # ====================================================
        # PROCESAR TODOS LOS CHATS
        # ====================================================

        resumen_resultados = []

        todos_los_detalles = []


        for archivo in archivos:

            agente = identificar_agente(
                archivo.name
            )


            # ------------------------------------------------
            # LEER TXT
            # ------------------------------------------------

            try:

                contenido = (
                    archivo
                    .getvalue()
                    .decode("utf-8")
                )

            except UnicodeDecodeError:

                contenido = (
                    archivo
                    .getvalue()
                    .decode(
                        "utf-8-sig",
                        errors="replace"
                    )
                )


            # ------------------------------------------------
            # LEER MENSAJES
            # ------------------------------------------------

            mensajes = leer_chat_whatsapp(
                contenido
            )


            # ------------------------------------------------
            # ANALIZAR MISMO PERÍODO
            # ------------------------------------------------

            df = analizar_solicitudes(
                mensajes,
                desde,
                hasta
            )


            # ------------------------------------------------
            # SI NO HAY REGISTROS
            # ------------------------------------------------

            if df.empty:

                total_solicitudes = 0

                total_contestadas = 0

                total_sin_contestar = 0

                porcentaje_respuesta = 0

                promedio_respuesta = None

                promedio_texto = "—"

                detalle = pd.DataFrame(
                    columns=[
                        "Fecha solicitud",
                        "Solicitante",
                        "RIF",
                        "Estado",
                        "Fecha respuesta",
                        "Tiempo respuesta (min)"
                    ]
                )


            # ------------------------------------------------
            # SI HAY REGISTROS
            # ------------------------------------------------

            else:

                detalle = relacionar_solicitudes(
                    df
                )


                total_solicitudes = len(
                    detalle
                )


                total_contestadas = (
                    detalle[
                        "Estado"
                    ]
                    .eq(
                        "Contestada"
                    )
                    .sum()
                )


                total_sin_contestar = (
                    detalle[
                        "Estado"
                    ]
                    .eq(
                        "Sin contestar"
                    )
                    .sum()
                )


                if total_solicitudes > 0:

                    porcentaje_respuesta = (

                        total_contestadas
                        / total_solicitudes
                        * 100
                    )

                else:

                    porcentaje_respuesta = 0


                tiempos_respuesta = detalle[
                    detalle[
                        "Estado"
                    ]
                    == "Contestada"
                ][
                    "Tiempo respuesta (min)"
                ]


                if not tiempos_respuesta.empty:

                    promedio_respuesta = (
                        tiempos_respuesta.mean()
                    )

                else:

                    promedio_respuesta = None


                promedio_texto = (
                    formatear_tiempo(
                        promedio_respuesta
                    )
                )


            # =================================================
            # RESUMEN DE ESTE AGENTE
            # =================================================

            resumen_resultados.append({

                "Agente":
                    agente,

                "Solicitudes":
                    total_solicitudes,

                "Contestados":
                    total_contestadas,

                "Sin Contestar":
                    total_sin_contestar,

                "Tasa de respuesta":
                    (
                        f"{porcentaje_respuesta:.1f}%"
                    ),

                "Tiempo promedio de respuesta":
                    promedio_texto
            })


            # =================================================
            # DETALLE DE ESTE AGENTE
            # =================================================

            if not detalle.empty:

                detalle = (
                    detalle.copy()
                )

                detalle.insert(
                    0,
                    "Agente",
                    agente
                )

                detalle[
                    "Observaciones"
                ] = ""

                todos_los_detalles.append(
                    detalle
                )


        # ====================================================
        # CREAR RESUMEN
        # ====================================================

        resumen_df = pd.DataFrame(
            resumen_resultados
        )


        # ====================================================
        # ORDENAR RESUMEN SEGÚN LISTA DE AGENTES
        # ====================================================

        orden_agentes = {
            agente: indice
            for indice, agente
            in enumerate(AGENTES)
        }


        resumen_df[
            "_orden"
        ] = resumen_df[
            "Agente"
        ].map(
            orden_agentes
        ).fillna(
            999
        )


        resumen_df = (
            resumen_df
            .sort_values(
                "_orden"
            )
            .drop(
                columns=[
                    "_orden"
                ]
            )
            .reset_index(
                drop=True
            )
        )


        # ====================================================
        # UNIR TODAS LAS SOLICITUDES
        # ====================================================

        if todos_los_detalles:

            detalle_general = pd.concat(
                todos_los_detalles,
                ignore_index=True
            )

        else:

            detalle_general = pd.DataFrame(
                columns=[
                    "Agente",
                    "Fecha solicitud",
                    "Solicitante",
                    "RIF",
                    "Estado",
                    "Fecha respuesta",
                    "Tiempo respuesta (min)",
                    "Observaciones"
                ]
            )


        # ====================================================
        # MOSTRAR ESTADÍSTICAS EN PANTALLA
        # ====================================================

        st.subheader(
            "Resumen por agente"
        )

        st.dataframe(
            resumen_df,
            use_container_width=True,
            hide_index=True
        )


        # ====================================================
        # TOTALES GENERALES
        # ====================================================

        solicitudes_generales = (
            resumen_df[
                "Solicitudes"
            ].sum()
        )

        contestadas_generales = (
            resumen_df[
                "Contestados"
            ].sum()
        )

        sin_contestar_generales = (
            resumen_df[
                "Sin Contestar"
            ].sum()
        )


        if solicitudes_generales > 0:

            tasa_general = (
                contestadas_generales
                / solicitudes_generales
                * 100
            )

        else:

            tasa_general = 0


        st.subheader(
            "Estadísticas generales"
        )


        met1, met2, met3, met4 = st.columns(
            4
        )


        with met1:

            st.metric(
                "Solicitudes",
                int(
                    solicitudes_generales
                )
            )


        with met2:

            st.metric(
                "Contestadas",
                int(
                    contestadas_generales
                )
            )


        with met3:

            st.metric(
                "Sin contestar",
                int(
                    sin_contestar_generales
                )
            )


        with met4:

            st.metric(
                "Tasa de respuesta",
                f"{tasa_general:.1f}%"
            )


        # ====================================================
        # GRÁFICA GENERAL
        # ====================================================

        st.subheader(
            "Solicitudes por agente"
        )


        grafica_agentes = (
            resumen_df[
                [
                    "Agente",
                    "Solicitudes",
                    "Contestados",
                    "Sin Contestar"
                ]
            ]
            .set_index(
                "Agente"
            )
        )


        st.bar_chart(
            grafica_agentes
        )


        # ====================================================
        # GENERAR EXCEL
        # ====================================================

        archivo_excel = crear_excel(
            resumen_df,
            detalle_general
        )


        st.divider()


        # ====================================================
        # DESCARGAR
        # ====================================================

        st.download_button(

            label=
                "📥 Descargar Excel",

            data=
                archivo_excel,

            file_name=(
                "estadisticas_whatsapp_"
                f"{fecha_desde.strftime('%d-%m-%Y')}"
                "al"
                f"{fecha_hasta.strftime('%d-%m-%Y')}"
                ".xlsx"
            ),

            mime=(
                "application/"
                "vnd.openxmlformats-officedocument."
                "spreadsheetml.sheet"
            )
        )